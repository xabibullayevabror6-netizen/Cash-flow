"""
Hisobotlarni Excel faylga chiqarish.

Nima uchun CSV emas, Excel:
  CFO hisobotni hamkasbiga yuboradi yoki arxivlaydi. CSV'da O'zbek/rus
  harflari Excel'da ko'pincha buziladi va raqamlar matnga aylanadi.
  XLSX bu muammolarni butunlay chetlab o'tadi.
"""
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.auth import get_current_user
from app.services.cash_flow_structure import classify, label as group_label
from app import models

router = APIRouter(prefix="/api/export", tags=["export"])

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FILL = PatternFill("solid", fgColor="0F1B24")
HEADER_FONT = Font(color="F3EEDF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13)
MONEY_FORMAT = "#,##0.00"


def _style_header(ws, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, widths: dict) -> None:
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def _filename(prefix: str) -> str:
    return f"{prefix}_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"


@router.get("/transactions")
def export_transactions(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    currency: Optional[str] = None,
    review_status: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """Operatsiyalar ro'yxati — filtrlar hisobga olinadi."""
    q = db.query(models.Transaction).filter(models.Transaction.company_id == user.company_id)
    if date_from:
        q = q.filter(models.Transaction.date >= date_from)
    if date_to:
        q = q.filter(models.Transaction.date <= date_to)
    if currency:
        q = q.filter(models.Transaction.currency == currency)
    if review_status:
        q = q.filter(models.Transaction.review_status == review_status)

    transactions = (
        q.options(joinedload(models.Transaction.category))
        .order_by(models.Transaction.date.desc(), models.Transaction.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Operatsiyalar"

    headers = [
        "Sana", "Yo'nalish", "Summa", "Valyuta", "Kontragent",
        "Kategoriya", "Provodka", "Qatlam", "Guruh",
        "Manba", "Holat", "To'lov tavsifi",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for t in transactions:
        direction = "in" if t.direction == models.Direction.in_ else "out"
        flow_type, group_key = classify(t.corr_account_code, direction)
        ws.append([
            t.date,
            "kirim" if direction == "in" else "chiqim",
            float(t.amount),
            t.currency,
            t.counterparty,
            t.category.name if t.category else "",
            t.corr_account_code or "",
            flow_type,
            group_label(group_key),
            t.category_source.value if t.category_source else "",
            t.review_status.value if t.review_status else "",
            t.raw_description or "",
        ])

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = MONEY_FORMAT

    _autosize(ws, {1: 12, 2: 9, 3: 18, 4: 9, 5: 38, 6: 24, 7: 10,
                   8: 12, 9: 28, 10: 10, 11: 14, 12: 46})

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer, media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{_filename("operatsiyalar")}"'},
    )


@router.get("/dashboard")
def export_dashboard(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    currency: Optional[str] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user),
):
    """
    Dashboard hisoboti: umumiy ko'rsatkichlar, xarajat tuzilmasi,
    kategoriya kesimi va kontragentlar — har biri alohida varaqda.
    """
    from app.routers.dashboard import (
        _conditions, _resolve_currency, cash_flow_structure, concentration,
    )

    active_currency, _av = _resolve_currency(db, user.company_id, currency)
    structure = cash_flow_structure(date_from, date_to, active_currency, db, user)
    risk = concentration("out", 20, date_from, date_to, active_currency, db, user)

    wb = Workbook()

    # --- 1. Umumiy ---
    ws = wb.active
    ws.title = "Umumiy"
    ws["A1"] = "Pul oqimi hisoboti"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Davr: {structure.period_start} — {structure.period_end}"
    ws["A3"] = f"Valyuta: {structure.currency or ''}"
    ws["A4"] = f"Yaratilgan: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    rows = [
        ("Ko'rsatkich", "Summa"),
        ("Asosiy faoliyat — tushum", structure.operating_in),
        ("Asosiy faoliyat — xarajat", structure.operating_out),
        ("Asosiy faoliyat — sof oqim", structure.operating_net),
        ("Investitsion faoliyat — sof", structure.investing_net),
        ("Moliyaviy faoliyat — sof", structure.financing_net),
        ("Sof o'zgarish", structure.net_change),
        ("Ichki ko'chirmalar (hisobga olinmagan)", structure.internal_volume),
    ]
    start = 6
    for offset, (label, value) in enumerate(rows):
        ws.cell(row=start + offset, column=1, value=label)
        ws.cell(row=start + offset, column=2, value=value)
    _style_header(ws, start, 2)
    for row in ws.iter_rows(min_row=start + 1, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _autosize(ws, {1: 42, 2: 20})

    # --- 2. Xarajat va tushum tuzilmasi ---
    ws = wb.create_sheet("Tuzilma")
    ws.append(["Qatlam", "Yo'nalish", "Guruh", "Summa", "Ulush", "Operatsiyalar"])
    _style_header(ws, 1, 6)
    for layer in structure.layers:
        for direction, groups in (("kirim", layer.inflow_groups), ("chiqim", layer.outflow_groups)):
            for group in groups:
                ws.append([layer.flow_type, direction, group.group,
                           group.amount, group.share, group.transaction_count])
    for row in ws.iter_rows(min_row=2):
        row[3].number_format = MONEY_FORMAT
        row[4].number_format = "0.0%"
    _autosize(ws, {1: 14, 2: 10, 3: 34, 4: 18, 5: 10, 6: 14})

    # --- 3. Kategoriya kesimi ---
    ws = wb.create_sheet("Kategoriyalar")
    ws.append(["Kategoriya", "Yo'nalish", "Summa"])
    _style_header(ws, 1, 3)

    conditions = _conditions(user.company_id, date_from, date_to, active_currency)
    from sqlalchemy import func as sa_func
    category_rows = (
        db.query(models.Category.name, models.Transaction.direction,
                 sa_func.sum(models.Transaction.amount))
        .outerjoin(models.Category, models.Category.id == models.Transaction.category_id)
        .filter(*conditions)
        .group_by(models.Category.name, models.Transaction.direction)
        .all()
    )
    for name, direction_value, amount in sorted(category_rows, key=lambda r: -float(r[2])):
        ws.append([
            name or "Kategoriyalanmagan",
            "kirim" if direction_value == models.Direction.in_ else "chiqim",
            float(amount),
        ])
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = MONEY_FORMAT
    _autosize(ws, {1: 34, 2: 10, 3: 18})

    # --- 4. Kontragentlar va konsentratsiya ---
    ws = wb.create_sheet("Kontragentlar")
    ws["A1"] = "Chiqim bo'yicha konsentratsiya"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Kontragentlar soni: {risk.counterparty_count}"
    ws["A3"] = f"Top-1 ulushi: {risk.top1_share:.1%}"
    ws["A4"] = f"Top-3 ulushi: {risk.top3_share:.1%}"
    ws["A5"] = f"Summaning 80% ini {risk.counterparties_for_80pct} ta kontragent tashkil qiladi"

    ws.append([])
    ws.append(["Kontragent", "Summa", "Ulush"])
    _style_header(ws, 7, 3)
    for item in risk.top_counterparties:
        ws.append([item.name, item.amount, item.share])
    for row in ws.iter_rows(min_row=8):
        row[1].number_format = MONEY_FORMAT
        row[2].number_format = "0.0%"
    _autosize(ws, {1: 44, 2: 18, 3: 10})

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer, media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{_filename("hisobot")}"'},
    )
